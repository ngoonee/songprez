# -*- coding: utf-8 -*-
#!/usr/bin/env python

from pprint import pprint
from .IO import opensong
import os
import openpyxl
import datetime

setpath = "/home/data/Dropbox/OpenSong/Sets/"
songpath = "/home/data/Dropbox/OpenSong/Songs/"
youthWL = ["jiamin", "joy", "egbert", "deborah", "josiah", "peiwen", "krystal", "samuel", "darren", "jasmine", "oonee", "melissa", "ernest", "sam", "brandon", "petre"]
adultWL = ["robin", "angie", "sukim", "chockseang", "angeline", "yow", "grace", "melissa", "suai", "chimeng", "ernest", "egbert", "sam", "deborah", "josiah"]
ccWL = ["oonee", "jane", "elizabeth", "elisabeth", "ashley", "caryn", "margaret", "natasha", "cynde", "adrienne", "jenna", "nicola", "oonhui"]


def cleanup_names():
    for l in opensong.list_files(setpath, recursive=True):
        for f in l:
            s = opensong.load_set(f)
            if os.path.basename(f) != s['set']['@name']:
                s['set']['@name'] = os.path.basename(f)
                opensong.write_set(f, s)
                

def check_worship_leader(name, list):
    name = ''.join(name.split()).lower().replace('-','')
    for n in list:
        if n in name:
            return True
    return False


def day_of_the_week(year, month, day):
    days = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")
    return days[datetime.date(int(year), int(month), int(day)).weekday()]


def write_excel():

    wb = openpyxl.Workbook()
    total = wb.get_active_sheet()
    total.title = "Historical Worship Lists"
    total.append(("Date", "Day of the week", "Worship leader/event"))
    youth = wb.create_sheet()
    youth.append(("Date", "Day of the week", "Worship leader/event"))
    youth.title = "Saturday Youth worship lists"
    saturdays = wb.create_sheet()
    saturdays.append(("Date", "Day of the week", "Worship leader/event"))
    saturdays.title = "Saturday other worship lists"
    cc = wb.create_sheet()
    cc.append(("Date", "Day of the week", "Worship leader/event"))
    cc.title = "Sunday CC worship lists"
    adult = wb.create_sheet()
    adult.append(("Date", "Day of the week", "Worship leader/event"))
    adult.title = "Sunday Adult worship lists"
    adultsongs = {}
    sundays = wb.create_sheet()
    sundays.append(("Date", "Day of the week", "Worship leader/event"))
    sundays.title = "Sunday other worship lists"
    for f in opensong.list_files(setpath, recursive=True):
        s = opensong.load_set(f)
        name = os.path.basename(f)
        day = day_of_the_week("20" + name[0:2], name[3:5], name[6:8])
        leader = name[9:]
        newrow = [name, day, leader]
        for song in s['set']['slide_groups']['slide_group']:
            newrow.append(song['@name'])
            if day == "Sunday" and check_worship_leader(leader, adultWL):
                adultsongs[song['@name']] = adultsongs.get(song['@name'], 0) + 1
        total.append(newrow)
        if day == "Saturday":
            if check_worship_leader(leader, youthWL):
                youth.append(newrow)
            else:
                saturdays.append(newrow)
        if day == "Sunday":
            if check_worship_leader(leader, ccWL):
                cc.append(newrow)
            elif check_worship_leader(leader, adultWL):
                adult.append(newrow)
            else:
                sundays.append(newrow)
    statistics = wb.create_sheet()
    statistics.append(("Song", "Frequency"))
    statistics.title = "Adult worship statistics"
    for s in adultsongs.keys():
        s2 = (s, adultsongs[s])
        statistics.append(s2)
    wb.save("/home/data/Dropbox/ChurchWork/WorshipTeam/SSMC Worship Ministry/AdultWorshipStatistics.xlsx")


if __name__=="__main__":
    write_excel()
